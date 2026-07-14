import math
from datetime import datetime
from io import BytesIO

from flask import Flask, render_template, request, send_file, make_response
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)


def calculate_annuity(principal, annual_rate, months):
    """
    Возвращает ежемесячный аннуитетный платёж.
    Если ставка 0, возвращает principal / months.
    """
    if annual_rate == 0:
        return principal / months
    monthly_rate = annual_rate / 12 / 100
    if monthly_rate == 0:
        return principal / months
    factor = (1 + monthly_rate) ** months
    return principal * (monthly_rate * factor) / (factor - 1)


def generate_schedule(principal, annual_rate, months, early_payments=None):
    """
    Генерирует график платежей с учётом досрочных погашений.
    early_payments: список словарей {'month': int, 'amount': float, 'type': 'payment' или 'term'}
    возвращает список словарей с данными по каждому месяцу.
    """
    if early_payments is None:
        early_payments = []

    # Сортируем досрочные платежи по месяцу
    early_payments = sorted(early_payments, key=lambda x: x['month'])

    schedule = []
    balance = principal
    monthly_rate = annual_rate / 12 / 100
    current_payment = calculate_annuity(principal, annual_rate, months)

    # Для отслеживания оставшегося срока
    remaining_months = months

    # Индекс для прохода по досрочным платежам
    ep_index = 0
    ep_count = len(early_payments)

    month = 1
    while balance > 0 and month <= remaining_months:
        # Проверяем, есть ли досрочный платёж в этом месяце
        early_extra = 0
        early_type = None
        if ep_index < ep_count and early_payments[ep_index]['month'] == month:
            early_extra = early_payments[ep_index]['amount']
            early_type = early_payments[ep_index]['type']
            ep_index += 1

        # Начисляем проценты за месяц
        interest = balance * monthly_rate
        # Основной долг в составе аннуитетного платежа
        principal_part = current_payment - interest
        if principal_part > balance:
            principal_part = balance
            current_payment = balance + interest  # последний платёж может быть меньше

        # Остаток после обычного платежа
        new_balance = balance - principal_part

        # Применяем досрочный платёж (если есть) – идёт на уменьшение основного долга
        if early_extra > 0:
            if new_balance > early_extra:
                new_balance -= early_extra
            else:
                early_extra = new_balance  # погашаем остаток
                new_balance = 0

        # Запись в график
        schedule.append({
            'month': month,
            'balance_start': round(balance, 2),
            'payment': round(current_payment, 2),
            'interest': round(interest, 2),
            'principal': round(principal_part, 2),
            'balance_end': round(new_balance, 2),
            'early_payment': round(early_extra, 2) if early_extra > 0 else None
        })

        balance = new_balance

        # Если остаток погашен – выходим
        if balance <= 0:
            break

        # Пересчёт после досрочного платежа (если был)
        if early_extra > 0 and early_type is not None:
            # Оставшийся срок
            remaining_months_after = remaining_months - month
            if remaining_months_after <= 0:
                break
            if early_type == 'payment':
                # Уменьшаем платёж, срок остаётся прежним
                current_payment = calculate_annuity(balance, annual_rate, remaining_months_after)
            elif early_type == 'term':
                # Срок уменьшается, платёж остаётся прежним
                # Оставляем current_payment без изменений
                # Но нужно пересчитать, сколько месяцев потребуется с текущим платежом
                # Для этого найдём минимальное количество месяцев, чтобы погасить balance
                # с текущим платежом и той же ставкой
                if annual_rate == 0:
                    new_months = math.ceil(balance / current_payment)
                else:
                    # Формула: n = -log(1 - (rate * balance) / payment) / log(1 + rate)
                    rate = monthly_rate
                    if rate == 0:
                        new_months = math.ceil(balance / current_payment)
                    else:
                        if current_payment <= balance * rate:
                            # Если платёж меньше процентов, никогда не погасится
                            # В реальности такое возможно только при ошибке, но мы просто урежем
                            new_months = remaining_months_after
                        else:
                            new_months = math.ceil(
                                -math.log(1 - (rate * balance) / current_payment) / math.log(1 + rate)
                            )
                remaining_months = month + new_months
                # Убедимся, что не превысим исходный срок
                if remaining_months > months:
                    remaining_months = months

        month += 1

    return schedule


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        try:
            # Получение данных из формы
            principal = float(request.form.get('principal', 0))
            months = int(request.form.get('months', 12))
            annual_rate = float(request.form.get('rate', 0))
            installment_plan = request.form.get('installment_plan') == 'on'

            # Если режим рассрочки, ставка принудительно 0
            if installment_plan:
                annual_rate = 0

            # Сбор досрочных платежей
            early_payments = []
            # Проверяем наличие полей early_month[], early_amount[], early_type[]
            early_months = request.form.getlist('early_month[]')
            early_amounts = request.form.getlist('early_amount[]')
            early_types = request.form.getlist('early_type[]')
            for m, a, t in zip(early_months, early_amounts, early_types):
                if m and a and t:
                    try:
                        month_val = int(m)
                        amount_val = float(a)
                        if month_val > 0 and amount_val > 0:
                            early_payments.append({
                                'month': month_val,
                                'amount': amount_val,
                                'type': t  # 'payment' или 'term'
                            })
                    except ValueError:
                        pass

            # Генерация графика
            schedule = generate_schedule(principal, annual_rate, months, early_payments)
            total_payment = sum(item['payment'] for item in schedule)
            total_interest = sum(item['interest'] for item in schedule)

            # Подготовка данных для шаблона
            context = {
                'schedule': schedule,
                'total_payment': round(total_payment, 2),
                'total_interest': round(total_interest, 2),
                'principal': principal,
                'months': months,
                'annual_rate': annual_rate,
                'installment_plan': installment_plan,
                'early_payments': early_payments,
            }
            return render_template('index.html', **context)

        except Exception as e:
            return render_template('index.html', error=str(e))

    # GET-запрос: показываем пустую форму
    return render_template('index.html', schedule=None)


@app.route('/download_excel')
def download_excel():
    """Генерирует и скачивает Excel-файл с графиком платежей."""
    # Параметры передаём через GET (для простоты)
    principal = float(request.args.get('principal', 0))
    months = int(request.args.get('months', 12))
    annual_rate = float(request.args.get('rate', 0))
    installment_plan = request.args.get('installment_plan') == 'true'
    if installment_plan:
        annual_rate = 0

    # Парсим досрочные платежи из строки
    early_payments = []
    early_data = request.args.get('early_payments', '')
    if early_data:
        import json
        try:
            early_payments = json.loads(early_data)
        except:
            pass

    schedule = generate_schedule(principal, annual_rate, months, early_payments)

    # Создаём Excel-книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "График платежей"

    # Заголовки
    headers = ['Месяц', 'Остаток на начало', 'Платёж', 'Проценты', 'Основной долг', 'Остаток на конец', 'Досрочный платёж']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2c1e16", end_color="2c1e16", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Заполнение данными
    for row_idx, item in enumerate(schedule, start=2):
        ws.cell(row=row_idx, column=1, value=item['month'])
        ws.cell(row=row_idx, column=2, value=item['balance_start'])
        ws.cell(row=row_idx, column=3, value=item['payment'])
        ws.cell(row=row_idx, column=4, value=item['interest'])
        ws.cell(row=row_idx, column=5, value=item['principal'])
        ws.cell(row=row_idx, column=6, value=item['balance_end'])
        ws.cell(row=row_idx, column=7, value=item['early_payment'] if item['early_payment'] else '')

    # Автоширина колонок
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Отправляем как файл
    return send_file(
        output,
        as_attachment=True,
        download_name=f"mortgage_schedule_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    app.run(debug=True)