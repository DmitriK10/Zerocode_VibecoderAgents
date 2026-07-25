import os
import random
from datetime import datetime
from openpyxl import Workbook, load_workbook
from config import Config

class ExcelBackup:
    def __init__(self):
        self.filepath = Config.EXCEL_FILE
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        if not os.path.exists(self.filepath):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            ws.append(["id", "full_name", "summ", "card_number", "birthday"])
            wb.save(self.filepath)
            print(f"📁 Создан новый Excel-файл: {self.filepath}")

    def save_user_data(self, full_name: str, birthday: str) -> bool:
        try:
            bday_obj = datetime.strptime(birthday, "%d.%m.%Y")
            bday_str = bday_obj.strftime("%Y-%m-%d")
            summ = round(random.uniform(100.0, 5000.0), 2)
            card_number = random.randint(1000000000000000, 9999999999999999)

            wb = load_workbook(self.filepath)
            ws = wb["Users"]

            last_id = 0
            if ws.max_row > 1:
                last_id = ws.cell(row=ws.max_row, column=1).value or 0
            new_id = last_id + 1

            ws.append([new_id, full_name, summ, card_number, bday_str])
            wb.save(self.filepath)
            print(f"📝 Данные добавлены в Excel для {full_name}")
            return True
        except Exception as e:
            print(f"❌ Ошибка сохранения в Excel: {e}")
            return False